import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Define styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill_navy = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
header_fill_teal = PatternFill(start_color="00A896", end_color="00A896", fill_type="solid")
header_fill_green = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# -------------------------------------------------------------
# TAB 1: Google_Ads_Data (Google Ads Tree Table Standard Format)
# -------------------------------------------------------------
ws_ads = wb.active
ws_ads.title = "Google_Ads_Data"

ads_headers = [
    "Day", "Campaign", "Account", "Ad group", "Search keyword", "Match type",
    "Quality score", "Exp. CTR", "Ad relevance", "Landing page exp.",
    "Impressions", "Clicks", "CTR", "Avg. CPC", "Cost", "Conversions", "Cost / conv.", "Conv. rate",
    "Search impr. share", "Search top IS", "Search abs. top IS", "Search lost IS (budget)", "Search lost IS (rank)",
    "Phone calls"
]

ws_ads.append(ads_headers)

sample_ads = [
    ["2026-08-15", "Alo_NMC_Search_Internal Medicine_Center_Samnan_Sun", "Sunny Clinics", "Internal Medicine - Exact", "internal medicine doctor samnan sharjah", "Exact", 9, "Above average", "Above average", "Above average", 310, 32, "10.32%", 8.40, 268.80, 6, 44.80, "18.75%", "88.0%", "88.0%", "62.0%", "8.0%", "4.0%", 3],
    ["2026-08-15", "Alo_NMC_Search_Cardiology_Royal_Khalifa_AUH", "AUH", "Cardiology - Exact High Intent", "best cardiologist in abu dhabi", "Exact", 9, "Above average", "Above average", "Above average", 320, 30, "9.38%", 14.20, 426.00, 5, 85.20, "16.67%", "84.0%", "82.5%", "54.0%", "12.0%", "5.5%", 2],
    ["2026-08-15", "Alo_NMC_Search_Orthopedics_Specialty_AlNahda_DXB", "DXB", "Orthopedics - Knee Replacement", "knee replacement surgeon dubai", "Exact", 9, "Above average", "Above average", "Above average", 280, 26, "9.28%", 22.40, 582.40, 5, 116.48, "19.23%", "86.0%", "85.0%", "59.0%", "9.0%", "5.0%", 2],
    ["2026-08-15", "Alo_NMC_Search_Cardiology_Royal_Sharjah_SHJ", "Northern Emirates", "Cardiology - Exact High Intent", "cardiologist in sharjah royal hospital", "Exact", 9, "Above average", "Above average", "Above average", 290, 28, "9.66%", 12.80, 358.40, 5, 71.68, "17.86%", "87.0%", "87.0%", "61.0%", "7.0%", "6.0%", 2],
    ["2026-08-15", "Alo_NMC_Search_Pediatrics_Center_Buhaira_Sun", "Sunny Clinics", "Pediatrics - Exact High Intent", "pediatrician buhaira corniche sharjah", "Exact", 9, "Above average", "Above average", "Above average", 360, 38, "10.56%", 7.50, 285.00, 8, 35.62, "21.05%", "90.0%", "90.0%", "69.0%", "5.0%", "5.0%", 4]
]

for row in sample_ads:
    ws_ads.append(row)

for col_idx in range(1, len(ads_headers) + 1):
    cell = ws_ads.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill_navy
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# -------------------------------------------------------------
# TAB 2: Call_Center_Leads
# -------------------------------------------------------------
ws_leads = wb.create_sheet(title="Call_Center_Leads")

leads_headers = [
    "ID", "Status", "Patient", "Phone", "Email", "Doctor", "Branch", "Department",
    "Lead priority", "Appointment Date", "Appointment Time", "Handled By", "Response Time", "Created At"
]

ws_leads.append(leads_headers)

sample_leads = [
    ["NMC-LD-10021", "Surgery Scheduled", "Rashid Al-Nuaimi", "+971 50 293 8492", "rashid.nuaimi@gmail.com", "Dr. Sanjay Sharma", "NMC Royal Hospital Khalifa City", "Cardiology", "High", "2026-08-18", "10:30", "Fatima Al-Mazrouei", "3 mins", "2026-08-15 09:15"],
    ["NMC-LD-10022", "Attended", "Mariam Al-Ali", "+971 52 481 9283", "mariam.ali@gmail.com", "Dr. Hisham Qasim", "Sunny Medical Centre Samnan", "Internal Medicine", "Urgent", "2026-08-16", "14:00", "Kareem Mansour", "2 mins", "2026-08-15 10:45"],
    ["NMC-LD-10023", "Booked", "Salem Al-Ketbi", "+971 55 938 1029", "salem.ketbi@gmail.com", "Dr. Monica Fakih", "NMC Royal Women's Hospital", "IVF & Fertility", "VIP", "2026-08-19", "11:15", "Rhea Chakraborty", "1 mins", "2026-08-15 11:30"],
    ["NMC-LD-10024", "Not Booked", "Rahul Verma", "+971 50 182 9384", "rahul.verma@gmail.com", "Dr. Omar Farooq", "Sunny Al Buhaira Medical Centre", "Pediatrics", "Low", "", "", "Zaid Al-Harbi", "28 mins", "2026-08-15 12:10"],
    ["NMC-LD-10025", "Booked", "Fatima Al-Kaabi", "+971 50 847 1928", "fatima.kaabi@gmail.com", "Dr. Philippe Macaire", "NMC Specialty Hospital Al Nahda", "Orthopedics", "High", "2026-08-20", "09:45", "Fatima Al-Mazrouei", "2 mins", "2026-08-15 13:00"]
]

for row in sample_leads:
    ws_leads.append(row)

for col_idx in range(1, len(leads_headers) + 1):
    cell = ws_leads.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill_teal
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# -------------------------------------------------------------
# TAB 3: Tree_Table_Setup_Guide
# -------------------------------------------------------------
ws_guide = wb.create_sheet(title="Tree_Table_Setup_Guide")
guide_title = ["Google Ads Report Editor Tree Table Configuration", ""]
ws_guide.append(guide_title)
ws_guide.append(["", ""])
ws_guide.append(["Row Hierarchy (First 4 Dimensions):", "1. Day  |  2. Campaign  |  3. Account  |  4. Ad group"])
ws_guide.append(["Metric Columns:", "Search keyword, Match type, Quality score, Exp. CTR, Ad relevance, Landing page exp., Impr., Clicks, CTR, Avg. CPC, Cost, Conv., Cost / conv., Conv. rate, Search impr. share, Search top IS, Search abs. top IS, Search lost IS (budget), Search lost IS (rank), Phone calls"])
ws_guide.append(["", ""])
ws_guide.append(["Dashboard Connection:", "Directly paste rows into Google_Ads_Data tab and click '🚀 NMC Sync > 🔄 Sync Data to Dashboard'."])

ws_guide.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True, color="107C41")

# Auto-adjust column widths
for ws in [ws_ads, ws_leads, ws_guide]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save("NMC_Hospital_Master_Template.xlsx")
print("Saved NMC_Hospital_Master_Template.xlsx with Google Ads Tree Table Schema!")
